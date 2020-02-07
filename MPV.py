# read xlsx
import openpyxl 
import csv

workbook = openpyxl.load_workbook("MVP01-6_Cars_by_make.xlsx")
sheet=workbook["MVP01-6(Cars by Make)"]

writerFileHandle = open("MVP_BMW_HONDA.csv", "w", newline='')
writer1 = csv.writer(writerFileHandle)

max_row = sheet.max_row
max_column = sheet.max_column
print("max row: " + str(max_row))
print("max column: " + str(max_column))

# Loop through every row to check if it is BMW or Honda
for i in range(1,max_row+1):
    
    #read cell
    make = sheet.cell(row=i, column=1).value
    if make == "B.M.W." or make == "HONDA":
        total = 0
        for col in range(2,12+1):
            total = total + sheet.cell(row=i, column=col).value
        row=[]
        row.append(make)
        row.append(total)
        #print(row)
        writer1.writerow(row)    
        print(make + " " + str(total))
        
# close csv
writerFileHandle.close()